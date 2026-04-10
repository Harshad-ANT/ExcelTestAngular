#!/usr/bin/env python3
"""
Convert BCS CAM Note Excel file to Univer IWorkbookData JSON template.
Includes: cell data, formulas, merged cells, column widths, row heights,
          borders, fills, fonts, alignment, data validations (dropdowns).
"""

import openpyxl
import json
import re
from openpyxl.utils import get_column_letter, column_index_from_string

EXCEL_FILE = 'BCS Cam Note Final - Standard - 30.06.25.xlsx'
OUTPUT_FILE = 'src/assets/bcs-cam-note-template.json'

INCLUDE_SHEETS = [
    '2. Proposal', 'Borrowers Details', 'Initial KYC', 'Completed Projects',
    'Upcoming Projects', 'Land Bank ', 'Unsold Stock & Leased Prop',
    'Group Entities', 'Other Businesses Details', 'Debt Details',
    'Sathbara & Mutation Entries', 'Projects Land Details',
    'FSI Calc & Approval Status', 'Sales MIS & Inventory',
    'Construction & Payment Sched', 'Project Factsheet', 'Ongoing Projects',
    'Complied Cost Working ',  # Hidden dependency for Compiled Cost
    'Compiled Cost Working ', 'Compiled Cost',
    'Collateral Security details ', 'Google Map', 'Result and Rating',
]

HIDDEN_SHEETS = {'Complied Cost Working '}

# Sheets with form-style layout (label + value columns) — dropdowns stay single-cell
FORM_SHEETS = {
    '2. Proposal',
    'Borrowers Details',
    'Project Factsheet',
    'Collateral Security details ',
}

# Cells containing "Dropdown" text — will be cleared (value handled by data validation)
DROPDOWN_MARKER_TEXTS = {'Dropdown', 'Dropdown ', 'dropdown', '(Dropdown )', '(Dopdown )'}

# Helper/instruction text patterns to remove (column D/E/F notes)
INSTRUCTION_PATTERNS = [
    r'^Auto fill link',
    r'^Link with',
    r'^Raw Data',
    r'^Value Will\s+come from',
    r'^Autofill',
    r'^Formula',
    r'^From Sale MIS',
    r'^Here, as per',
    r'^These details shall be exported',
    r'^Update List',
    r'^Add other\s+Professionals',
    r'^This will be taken from',
    r'^This Table will be autofill',
    r'^Taken Form',
    r'^Completed project madhil',
    r'^Consider average rate',
    r'^Note :',
    r'^Note:',
    r'^Aaplyala',
    r'^Vijaylaxmi|^Lotewadi|^Akshay Mane|^Umbergaon|^kharsundi',
    r'^Ground / Stilt',
    r'^Shops, Offices',
    r'^Wing & No of Tower',
    r'^Pls make Per ft Rate',
    r'^As per initial KYC',
]

# Build Univer style index system
style_counter = 0
style_map = {}  # key -> style_id
styles_dict = {}  # style_id -> style definition


def get_color_rgb(color_obj):
    """Extract hex RGB from openpyxl color, return None if invalid."""
    if not color_obj:
        return None
    try:
        if color_obj.type == 'rgb' and color_obj.rgb:
            rgb = str(color_obj.rgb)
            if rgb.startswith('FF') and len(rgb) == 8 and rgb != '00000000':
                return '#' + rgb[2:]  # Strip alpha, add #
            elif rgb.startswith('00') and rgb == '00000000':
                return None
        elif color_obj.type == 'theme':
            # Map common theme colors
            theme_map = {
                0: '#FFFFFF', 1: '#000000', 2: '#E7E6E6', 3: '#44546A',
                4: '#4472C4', 5: '#ED7D31', 6: '#A5A5A5', 7: '#FFC000',
                8: '#5B9BD5', 9: '#70AD47',
            }
            return theme_map.get(color_obj.theme)
    except:
        pass
    return None


def border_style_to_univer(style):
    """Convert Excel border style to Univer border style number."""
    mapping = {
        'thin': 1,
        'medium': 2,
        'thick': 3,
        'dashed': 5,
        'dotted': 4,
        'double': 6,
        'hair': 7,
        'mediumDashed': 8,
        'dashDot': 9,
        'mediumDashDot': 10,
        'dashDotDot': 11,
        'mediumDashDotDot': 12,
        'slantDashDot': 13,
    }
    return mapping.get(style, 1)


def make_border_obj(side):
    """Create Univer border side object from openpyxl side."""
    if not side or not side.style:
        return None
    color = get_color_rgb(side.color) or '#000000'
    return {
        's': border_style_to_univer(side.style),
        'cl': {'rgb': color},
    }


def build_cell_style(cell):
    """Build a Univer style dict from cell formatting."""
    s = {}

    # Background fill
    try:
        if cell.fill and cell.fill.patternType == 'solid':
            bg = get_color_rgb(cell.fill.fgColor)
            if bg:
                s['bg'] = {'rgb': bg}
    except:
        pass

    # Font
    try:
        if cell.font:
            if cell.font.bold:
                s['bl'] = 1  # bold
            if cell.font.italic:
                s['it'] = 1
            if cell.font.size:
                s['fs'] = cell.font.size
            fc = get_color_rgb(cell.font.color)
            if fc and fc != '#000000':
                s['cl'] = {'rgb': fc}
            if cell.font.name:
                s['ff'] = cell.font.name
    except:
        pass

    # Borders
    try:
        if cell.border:
            bd = {}
            for side_name, prop_name in [('t', 'top'), ('b', 'bottom'), ('l', 'left'), ('r', 'right')]:
                side = getattr(cell.border, prop_name)
                bo = make_border_obj(side)
                if bo:
                    bd[side_name] = bo
            if bd:
                s['bd'] = bd
    except:
        pass

    # Alignment
    try:
        if cell.alignment:
            ha_map = {'left': 0, 'center': 1, 'right': 2, 'general': 0, 'justify': 4, 'fill': 5}
            va_map = {'top': 0, 'center': 1, 'bottom': 2}
            if cell.alignment.horizontal and cell.alignment.horizontal in ha_map:
                s['ht'] = ha_map[cell.alignment.horizontal]
            if cell.alignment.vertical and cell.alignment.vertical in va_map:
                s['vt'] = va_map[cell.alignment.vertical]
            if cell.alignment.wrap_text:
                s['tb'] = 3  # wrap text
    except:
        pass

    return s if s else None


def get_or_create_style_id(style_dict):
    """Get existing style ID or create new one."""
    global style_counter
    if not style_dict:
        return None

    key = json.dumps(style_dict, sort_keys=True)
    if key in style_map:
        return style_map[key]

    style_id = str(style_counter)
    style_counter += 1
    style_map[key] = style_id
    styles_dict[style_id] = style_dict
    return style_id


def is_instruction_text(val):
    """Check if cell value is an instruction/helper text to be removed."""
    if not isinstance(val, str):
        return False
    for pattern in INSTRUCTION_PATTERNS:
        if re.match(pattern, val.strip(), re.IGNORECASE):
            return True
    return False


def is_dropdown_marker(val):
    """Check if cell value is a 'Dropdown' marker text."""
    if not isinstance(val, str):
        return False
    return val.strip().rstrip(')').lstrip('(').strip() in DROPDOWN_MARKER_TEXTS or val.strip() in DROPDOWN_MARKER_TEXTS


def has_significant_style(cell):
    """Check if an empty cell has borders or fills worth preserving."""
    try:
        if cell.fill and cell.fill.patternType == 'solid':
            bg = get_color_rgb(cell.fill.fgColor)
            if bg:
                return True
        if cell.border:
            for side in [cell.border.top, cell.border.bottom, cell.border.left, cell.border.right]:
                if side and side.style:
                    return True
    except:
        pass
    return False


def convert_cell_value(cell):
    """Convert a cell value to Univer cell data format."""
    val = cell.value
    if val is None:
        # Only include empty cells with borders or fills
        if has_significant_style(cell):
            style = build_cell_style(cell)
            sid = get_or_create_style_id(style)
            if sid is not None:
                return {'s': sid}
        return None

    # Skip instruction text
    if is_instruction_text(val):
        style = build_cell_style(cell)
        sid = get_or_create_style_id(style)
        if sid is not None:
            return {'s': sid}
        return None

    # Skip dropdown marker text
    if is_dropdown_marker(val):
        style = build_cell_style(cell)
        sid = get_or_create_style_id(style)
        if sid is not None:
            return {'s': sid}
        return None

    cell_data = {}

    # Apply style
    style = build_cell_style(cell)
    sid = get_or_create_style_id(style)
    if sid is not None:
        cell_data['s'] = sid

    if isinstance(val, str) and val.startswith('='):
        formula = val
        if 'DATEDIF' in formula:
            match = re.search(r"DATEDIF\((.+?),(.+?),", formula)
            if match:
                start = match.group(1).strip()
                end = match.group(2).strip()
                formula = f'=({end}-{start})/30'
        formula = re.sub(r'SUBTOTAL\(\s*9\s*,', 'SUM(', formula)
        if formula.startswith('=+'):
            formula = '=' + formula[2:]
        formula = formula.replace('=++', '=').replace('++', '+')
        cell_data['f'] = formula
        cell_data['si'] = None
    elif isinstance(val, bool):
        cell_data['v'] = 1 if val else 0
        cell_data['t'] = 2
    elif isinstance(val, (int, float)):
        cell_data['v'] = val
        cell_data['t'] = 2
    elif isinstance(val, str):
        cleaned = val.strip()
        if cleaned == '\xa0' or cleaned == '':
            if 's' in cell_data:
                return cell_data
            return None
        cell_data['v'] = val
        cell_data['t'] = 1
    else:
        try:
            cell_data['v'] = str(val)
            cell_data['t'] = 1
        except:
            return None

    return cell_data


def parse_range_string(sqref_str):
    """Parse Excel range string like 'C8' or 'C18:C19' to list of (row, col) tuples."""
    ranges = []
    for part in str(sqref_str).split():
        if ':' in part:
            start, end = part.split(':')
            start_col = column_index_from_string(re.match(r'([A-Z]+)', start).group(1)) - 1
            start_row = int(re.search(r'(\d+)', start).group(1)) - 1
            end_col = column_index_from_string(re.match(r'([A-Z]+)', end).group(1)) - 1
            end_row = int(re.search(r'(\d+)', end).group(1)) - 1
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    ranges.append((r, c))
        else:
            col = column_index_from_string(re.match(r'([A-Z]+)', part).group(1)) - 1
            row = int(re.search(r'(\d+)', part).group(1)) - 1
            ranges.append((row, col))
    return ranges


def convert_data_validations(ws):
    """Convert Excel data validations + dropdown indicators to Univer format."""
    validations = []

    # 1. Explicit Excel data validations — extend single-cell DVs to table column
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list':
                continue
            formula = dv.formula1 or ''
            formula = formula.strip('"')
            options = [v.strip() for v in formula.split(',') if v.strip()]
            if not options:
                continue

            sqref_parts = str(dv.sqref).split()
            expanded_cells = []

            for part in sqref_parts:
                if ':' in part:
                    # Already a range — use as-is
                    expanded_cells.extend(parse_range_string(part))
                else:
                    # Single cell — extend down the table column
                    col_idx = column_index_from_string(
                        re.match(r'([A-Z]+)', part).group(1)) - 1
                    start_row = int(re.search(r'(\d+)', part).group(1)) - 1
                    last_data_row = start_row

                    # FORM sheets: single cell only. TABLE sheets: extend column.
                    is_form = ws.title in FORM_SHEETS

                    if is_form:
                        # Form layout — single cell only
                        expanded_cells.append((start_row, col_idx))
                    else:
                        # Table layout — extend to full column
                        check_cols = list(range(0, min(3, ws.max_column)))
                        check_cols.extend(range(max(0, col_idx - 1), min(col_idx + 2, ws.max_column)))
                        check_cols = sorted(set(check_cols))

                        def row_has_content(row_1based):
                            for cc in check_cols:
                                cell_check = ws.cell(row=row_1based, column=cc + 1)
                                try:
                                    v = cell_check.value
                                    if v is not None and str(v).strip():
                                        return True
                                except:
                                    continue
                            return False

                        # Look UPWARD to find first data row
                        first_data_row = start_row
                        for r in range(start_row - 1, max(0, start_row - 20), -1):
                            if row_has_content(r + 1):
                                hcell = ws.cell(row=r + 1, column=col_idx + 1)
                                hv = hcell.value
                                if hv and isinstance(hv, str) and hv.strip().isupper():
                                    first_data_row = r + 1
                                    break
                                first_data_row = r
                            else:
                                break

                        # Look DOWNWARD to find last data row
                        last_data_row = start_row
                        for r in range(start_row, min(start_row + 100, ws.max_row)):
                            if row_has_content(r + 1):
                                last_data_row = r
                            elif r + 1 < ws.max_row and row_has_content(r + 2):
                                continue
                            else:
                                break

                        # Empty template — extend 10 rows
                        if last_data_row == start_row:
                            last_data_row = start_row + 10

                        for r in range(first_data_row, last_data_row + 1):
                            expanded_cells.append((r, col_idx))

                    for r in range(start_row, last_data_row + 1):
                        expanded_cells.append((r, col_idx))

            for row, col in expanded_cells:
                validations.append({
                    'row': row,
                    'col': col,
                    'options': options,
                })

    # 2. Implicit dropdowns from "Dropdown" / "Drop down" marker text
    dropdown_cells = set()
    for v in validations:
        dropdown_cells.add((v['row'], v['col']))

    def parse_options(text):
        """Parse option values from a text string."""
        text = text.strip()
        text = re.sub(r'^\xa0', '', text)  # Remove leading NBSP
        text = re.sub(r'^\(|\)$', '', text).strip()  # Remove wrapping parens
        text = re.sub(r'\(?\s*[Dd]o?p?down[:\s]*\)?', '', text).strip()
        text = text.rstrip(',').strip()
        if not text:
            return []
        if '/' in text and ',' not in text:
            options = [o.strip() for o in text.split('/') if o.strip()]
        elif ',' in text:
            options = [o.strip() for o in text.split(',') if o.strip()]
        else:
            return []
        # Clean individual options
        options = [re.sub(r'\(?\s*[Dd]o?p?down[:\s]*\)?', '', o).strip() for o in options]
        return [o for o in options if o]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            try:
                _ = cell.column_letter
            except AttributeError:
                continue
            if not cell.value or not isinstance(cell.value, str):
                continue
            val = cell.value.strip()
            is_dropdown_marker = bool(re.match(r'^[Dd]rop\s*down', val.replace('\xa0', ' ')))
            if not is_dropdown_marker:
                continue

            # Determine target_row, target_col, and options
            target_row = None
            target_col = None
            options = None

            colon_idx = val.find(':')
            if colon_idx > 0:
                # Marker contains inline options (e.g. "Drop down: Yes /NO")
                inline = val[colon_idx + 1:]
                options = parse_options(inline)
                if len(options) >= 2:
                    target_row = cell.row - 1
                    target_col = cell.column - 2
            else:
                # Standard pattern: options are in the cell to the LEFT
                if cell.column > 1:
                    target_cell = ws.cell(row=cell.row, column=cell.column - 1)
                    if target_cell.value and isinstance(target_cell.value, str):
                        options = parse_options(target_cell.value)
                        target_row = cell.row - 1
                        target_col = cell.column - 2

            if target_row is None or target_col is None or target_col < 0:
                continue
            if not options or len(options) < 2:
                continue
            if (target_row, target_col) in dropdown_cells:
                continue

            is_form = ws.title in FORM_SHEETS

            if is_form:
                # Form layout — single cell only
                if (target_row, target_col) not in dropdown_cells:
                    validations.append({'row': target_row, 'col': target_col, 'options': options})
                    dropdown_cells.add((target_row, target_col))
            else:
                # Table layout — extend to full column
                check_cols_impl = list(range(0, min(3, ws.max_column)))
                check_cols_impl.extend(range(max(0, target_col - 1), min(target_col + 2, ws.max_column)))
                check_cols_impl = sorted(set(check_cols_impl))

                def impl_row_has_content(row_1based):
                    for cc in check_cols_impl:
                        c_check = ws.cell(row=row_1based, column=cc + 1)
                        try:
                            v = c_check.value
                            if v is not None and str(v).strip():
                                return True
                        except:
                            continue
                    return False

                first_r = target_row
                for r in range(target_row - 1, max(0, target_row - 20), -1):
                    if impl_row_has_content(r + 1):
                        hcell = ws.cell(row=r + 1, column=target_col + 1)
                        hv = hcell.value
                        if hv and isinstance(hv, str) and hv.strip().isupper():
                            first_r = r + 1
                            break
                        first_r = r
                    else:
                        break

                last_r = target_row
                for r in range(target_row, min(target_row + 100, ws.max_row)):
                    if impl_row_has_content(r + 1):
                        last_r = r
                    elif r + 1 < ws.max_row and impl_row_has_content(r + 2):
                        continue
                    else:
                        break

                if last_r == target_row:
                    last_r = target_row + 10

                for r in range(first_r, last_r + 1):
                    if (r, target_col) not in dropdown_cells:
                        validations.append({'row': r, 'col': target_col, 'options': options})
                        dropdown_cells.add((r, target_col))

    # Deduplicate
    seen = set()
    unique_validations = []
    for v in validations:
        key = (v['row'], v['col'])
        if key not in seen:
            seen.add(key)
            unique_validations.append(v)

    return unique_validations


def convert_sheet(ws, sheet_index):
    """Convert an openpyxl worksheet to Univer sheet data format."""
    sheet_id = f'sheet-{sheet_index:02d}'
    sheet_name = ws.title

    max_row = max(ws.max_row, 100)
    max_col = max(ws.max_column, 26)

    # Build cellData with styles
    cell_data = {}

    # First pass: cells with borders (including empty cells in border regions)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            try:
                # Skip merged cells (they throw AttributeError)
                _ = cell.column_letter
            except AttributeError:
                continue

            converted = convert_cell_value(cell)
            if converted is not None:
                row_idx = cell.row - 1
                col_idx = cell.column - 1
                if row_idx not in cell_data:
                    cell_data[row_idx] = {}
                cell_data[row_idx][col_idx] = converted

    # Build mergeData
    merge_data = []
    for mc in ws.merged_cells.ranges:
        merge_data.append({
            'startRow': mc.min_row - 1,
            'startColumn': mc.min_col - 1,
            'endRow': mc.max_row - 1,
            'endColumn': mc.max_col - 1,
        })

    # Build columnData (widths)
    column_data = {}
    for col_idx in range(max_col):
        col_letter = get_column_letter(col_idx + 1)
        dim = ws.column_dimensions.get(col_letter)
        if dim and dim.width:
            pixel_width = int(dim.width * 7.5)
            column_data[col_idx] = {'w': max(pixel_width, 50)}

    # Build rowData (heights) — add padding
    row_data = {}
    for row_idx in range(ws.max_row):
        dim = ws.row_dimensions.get(row_idx + 1)
        if dim and dim.height:
            row_data[row_idx] = {'h': max(dim.height + 4, 24)}
        else:
            row_data[row_idx] = {'h': 24}  # Minimum height for readability

    # Data validations (dropdowns)
    validations = convert_data_validations(ws)

    sheet_config = {
        'id': sheet_id,
        'name': sheet_name,
        'rowCount': max_row,
        'columnCount': max_col,
        'cellData': cell_data,
        'mergeData': merge_data,
        'columnData': column_data,
        'rowData': row_data,
        'defaultColumnWidth': 100,
        'defaultRowHeight': 24,
        'hidden': 1 if sheet_name in HIDDEN_SHEETS else 0,
    }

    if validations:
        sheet_config['_validations'] = validations

    return sheet_id, sheet_config


def main():
    global style_counter, style_map, styles_dict
    style_counter = 0
    style_map = {}
    styles_dict = {}

    print(f'Loading {EXCEL_FILE}...')
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

    sheets = {}
    sheet_order = []
    all_validations = {}
    sheet_index = 1

    for sheet_name in wb.sheetnames:
        if sheet_name not in INCLUDE_SHEETS:
            continue

        print(f'  Converting: {sheet_name} ({sheet_index}/{len(INCLUDE_SHEETS)})')
        ws = wb[sheet_name]
        sheet_id, sheet_config = convert_sheet(ws, sheet_index)

        # Extract validations separately
        if '_validations' in sheet_config:
            all_validations[sheet_id] = sheet_config.pop('_validations')

        sheets[sheet_id] = sheet_config
        sheet_order.append(sheet_id)
        sheet_index += 1

    # Build IWorkbookData
    workbook_data = {
        'id': 'bcs-cam-note-template',
        'name': 'BCS CAM Note',
        'appVersion': '0.1.0',
        'locale': 'EN_US',
        'styles': styles_dict,
        'sheetOrder': sheet_order,
        'sheets': sheets,
    }

    # Embed data validations in resources (Univer plugin resource format)
    if all_validations:
        import uuid
        dv_resources = {}
        for sheet_id, dv_list in all_validations.items():
            rules = []
            for dv in dv_list:
                rule = {
                    'uid': str(uuid.uuid4())[:8],
                    'type': 'list',
                    'formula1': ','.join(dv['options']),
                    'allowBlank': True,
                    'showDropDown': True,
                    'ranges': [{
                        'startRow': dv['row'],
                        'endRow': dv['row'],
                        'startColumn': dv['col'],
                        'endColumn': dv['col'],
                    }],
                }
                rules.append(rule)
            dv_resources[sheet_id] = rules

        workbook_data['resources'] = [{
            'name': 'SHEET_DATA_VALIDATION_PLUGIN',
            'data': json.dumps(dv_resources),
        }]

    print(f'\nWriting {OUTPUT_FILE}...')
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(workbook_data, f, ensure_ascii=False, indent=None)

    # Stats
    total_formulas = 0
    total_cells = 0
    total_merges = 0
    total_styled = 0
    for sid, s in sheets.items():
        total_merges += len(s['mergeData'])
        for row_data in s['cellData'].values():
            for cell in row_data.values():
                total_cells += 1
                if 'f' in cell:
                    total_formulas += 1
                if 's' in cell:
                    total_styled += 1

    file_size = len(json.dumps(workbook_data, ensure_ascii=False))
    print(f'\n=== STATS ===')
    print(f'Sheets: {len(sheets)}')
    print(f'Total cells: {total_cells}')
    print(f'Total formulas: {total_formulas}')
    print(f'Total merged regions: {total_merges}')
    print(f'Total styled cells: {total_styled}')
    print(f'Unique styles: {len(styles_dict)}')
    print(f'Data validations: {sum(len(v) for v in all_validations.values())}')
    print(f'JSON size: {file_size / 1024:.1f} KB')
    print('Done!')


if __name__ == '__main__':
    main()
